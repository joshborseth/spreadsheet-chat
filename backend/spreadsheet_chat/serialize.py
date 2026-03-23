from __future__ import annotations

from openpyxl import load_workbook


def workbook_to_llm_text(path: str) -> str:
    """
    Dump every non-empty cell as ``SheetName!A1 = repr(value)``.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for sheet in wb.worksheets:
            lines.append(f"### Sheet: {sheet.title}")
            try:
                dim = sheet.calculate_dimension()
            except Exception:  # noqa: BLE001
                dim = "?"
            lines.append(f"Used range: {dim}")
            for row in sheet.iter_rows():
                for cell in row:
                    val = cell.value
                    if val is None:
                        continue
                    if val == "":
                        continue
                    coord = f"{sheet.title}!{cell.coordinate}"
                    lines.append(f"{coord} = {val!r}")
            lines.append("")
    finally:
        wb.close()
    return "\n".join(lines).strip()
