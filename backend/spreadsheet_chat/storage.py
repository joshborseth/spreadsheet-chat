from __future__ import annotations

import tempfile
import uuid
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class StoredFile:
    path: Path
    filename: str
    sheet_names: list[str]


_store: dict[str, StoredFile] = {}
_upload_dir = Path(tempfile.mkdtemp(prefix="spreadsheet_chat_uploads_"))


def save_upload(*, content: bytes, filename: str) -> tuple[str, StoredFile]:
    if not content.startswith(b"PK"):
        raise ValueError("Not a valid .xlsx (ZIP) file")
    file_id = str(uuid.uuid4())
    dest = _upload_dir / f"{file_id}.xlsx"
    dest.write_bytes(content)
    wb = load_workbook(dest, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
    finally:
        wb.close()
    rec = StoredFile(path=dest, filename=filename, sheet_names=sheet_names)
    _store[file_id] = rec
    return file_id, rec


def get(file_id: str) -> StoredFile | None:
    return _store.get(file_id)
